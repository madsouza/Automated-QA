import requests
import datetime
import openpyxl
import pyglet
from pyglet.gl import gl
import authenticate

background_color = (0.98, 0.98, 0.98, 1)
proceed = False


class AuthenticationWindow(pyglet.window.Window):
    def __init__(self, *args, **kwargs):
        self.success = False
        super(AuthenticationWindow, self).__init__(*args, **kwargs)
        self.batch = pyglet.graphics.Batch()
        self.width = AuthenticationWindow.width.__get__(self)
        self.height = AuthenticationWindow.height.__get__(self)
        self.labels = pyglet.graphics.Batch()
        self.spr = pyglet.graphics.Batch()
        self.background_color = (0.98, 0.98, 0.98, 1)
        self.instruction_label = pyglet.text.Label('Please log in using QA track login', font_name='Helvetica',
                                                   font_size=12, x=100, y=175, color=(0, 147, 199, 255),
                                                   bold=True, batch=self.labels)
        self.user_text = pyglet.text.Label('Username', font_name='Helvetica', font_size=12, x=100, y=135,
                                           color=(0, 147, 199, 255), bold=True, batch=self.labels)

        self.incorrect_text = pyglet.text.Label('', font_name='Helvetica', font_size=10,
                                                x=140, y=55, color=(255, 0, 0, 255), bold=True, batch=self.labels)

        self.pass_text = pyglet.text.Label('Password', font_size=12, x=100, y=85,
                                           color=(0, 147, 199, 255), bold=True, batch=self.labels)

        self.vertex_list = pyglet.graphics.vertex_list(16, ('v2f', (
            # User name
            # horizontal lines
            200,
            128,
            400,
            128,

            200,
            150,
            400,
            150,

            # vertical lines
            200,
            128,
            200,
            150,

            400,
            128,
            400,
            150,

            # password
            # Horizontal lines
            200,
            100,
            400,
            100,

            200,
            78,
            400,
            78,

            # Vertical lines
            200,
            100,
            200,
            78,

            400,
            100,
            400,
            78,
            )),
                                                       ('c3B', (0, 0, 200) * 16))

        white = pyglet.image.load('grey.png')
        white.width = int(200)
        white.height = int(22)
        green = pyglet.image.load('green.png')
        green.width = int(200)
        green.height = int(30)

        green2 = pyglet.image.load('green_2.jpg')
        green2.width = int(200)
        green2.height = int(30)
        # self.user_sprite = pyglet.sprite.Sprite(white, x=200, y=128, batch=self.spr)
        # self.user_sprite.visible = False
        self.user_input = pyglet.text.Label('', font_name='Helvetica', font_size=12, x=205, y=135,
                                            color=(0, 147, 199, 255),bold=True)

        # self.pass_sprite = pyglet.sprite.Sprite(white, x=200, y=78, batch=self.spr)
        # self.pass_sprite.visible = True

        self.widgets = [
            TextWidget('user', 205, 130, self.width - 210, self.batch),
            TextWidget('password', 205, 80, self.width - 210, self.batch)
        ]
        self.text_cursor = self.get_system_mouse_cursor('text')

        self.focus = None
        self.set_focus(self.widgets[0])

        self.confirm_button = pyglet.sprite.Sprite(green, x=150, y=20, batch=self.spr)
        self.confirm_button_active = pyglet.sprite.Sprite(green2, x=150, y=20, batch=self.spr)
        self.confirm_button_active.visible = False
        self.confirm_text = pyglet.text.Label('Log In', font_name='Helvetica', font_size=12, x=225, y=28,
                                              color=(0, 147, 199, 255), bold=True, batch=self.labels)

        self.user = None
        self.password = None

    def on_draw(self):
        gl.glClearColor(*background_color)
        AuthenticationWindow.clear(self)
        self.spr.draw()
        self.user_input.draw()
        self.vertex_list.draw(pyglet.gl.GL_LINES)
        self.batch.draw()
        self.labels.draw()

    def on_mouse_motion(self, x, y, dx, dy):
        if 150 < x < 350 and 20 < y < 50:
            self.confirm_button_active.visible = True
            self.confirm_button.visible = False
        else:
            self.confirm_button_active.visible = False
            self.confirm_button.visible = True
        for widget in self.widgets:
            if widget.hit_test(x, y):
                self.set_mouse_cursor(self.text_cursor)
                break
        else:
            self.set_mouse_cursor(None)

    def on_mouse_press(self, x, y, button, modifiers):
        if 150 < x < 350 and 20 < y < 50:
            self.user = self.widgets[0].document.text
            self.password = self.widgets[1].document.text
            self.confirm()

        for widget in self.widgets:
            if widget.hit_test(x, y):
                self.set_focus(widget)
                break
        else:
            self.set_focus(None)

        if self.focus:
            self.focus.caret.on_mouse_press(x, y, button, modifiers)

    def on_mouse_drag(self, x, y, dx, dy, buttons, modifiers):
        if self.focus:
            self.focus.caret.on_mouse_drag(x, y, dx, dy, buttons, modifiers)

    def on_text(self, text):
        if self.focus:
            self.focus.caret.on_text(text)

    def on_text_motion(self, motion):
        if self.focus:
            self.focus.caret.on_text_motion(motion)

    def on_text_motion_select(self, motion):
        if self.focus:
            self.focus.caret.on_text_motion_select(motion)

    def on_key_press(self, symbol, modifiers):
        if symbol == pyglet.window.key.TAB:
            if modifiers & pyglet.window.key.MOD_SHIFT:
                dir = -1
            else:
                dir = 1

            if self.focus in self.widgets:
                i = self.widgets.index(self.focus)
            else:
                i = 0
                dir = 0

            self.set_focus(self.widgets[(i + dir) % len(self.widgets)])

        elif symbol == pyglet.window.key.ESCAPE:
            pyglet.app.exit()

    def set_focus(self, focus):
        if self.focus:
            self.focus.caret.visible = False
            self.focus.caret.mark = self.focus.caret.position = 0

        self.focus = focus
        if self.focus:
            self.focus.caret.visible = True
            self.focus.caret.mark = 0
            self.focus.caret.position = len(self.focus.document.text)

    def confirm(self):
        if authenticate.check_user(self.user, self.password):
            self.success = True
            self.close()
        else:
            self.incorrect_text.text = 'Incorrect username or password'

    def return_info(self):
        return self.user, self.password, self.success



class Rectangle(object):
    '''Draws a rectangle into a batch.'''

    def __init__(self, x1, y1, x2, y2, batch):
        self.vertex_list = batch.add(4, pyglet.gl.GL_QUADS, None,
                                     ('v2i', [x1, y1, x2, y1, x2, y2, x1, y2]),
                                     ('c4B', [200, 200, 220, 255] * 4)
                                     )


class TextWidget(object):
    def __init__(self, text, x, y, width, batch):
        self.document = pyglet.text.document.UnformattedDocument(text)
        self.document.set_style(0, len(self.document.text),
                                dict(color=(0, 0, 0, 255))
                                )
        font = self.document.get_font()
        height = font.ascent - font.descent

        self.layout = pyglet.text.layout.IncrementalTextLayout(
            self.document, width, height, multiline=False, batch=batch)
        self.caret = pyglet.text.caret.Caret(self.layout)

        self.layout.x = x
        self.layout.y = y

        # Rectangular outline
        pad = 2
        self.rectangle = Rectangle(x - pad, y - pad,
                                   x + width + pad, y + height + pad, batch)

    def hit_test(self, x, y):
        return (0 < x - self.layout.x < self.layout.width and
                0 < y - self.layout.y < self.layout.height)


def read_file():
    with open('user_info', 'r') as f:
        info = f.read().splitlines()
    ip = info[0]
    # username = info[1][6:]
    # passwrd = info[1][10:]
    return ip


def login(user, password):
    root = ip_address
    login_url = root + "accounts/login/"
    s = requests.Session()

    # we need to GET the login page so we can retrieve the csrf token
    s.get(login_url)
    token = s.cookies['csrftoken']

    login_data = {
        'username': user,
        'password': password,
        'csrfmiddlewaretoken': token
    }
    # perform the login
    login_resp = s.post(login_url, data=login_data)
    token = s.cookies['csrftoken']
    print token
    return root, token, s


def test_number(machine, test):
    wb = openpyxl.load_workbook('tests.xlsx')
    ws = wb['test_urls']
    row = filter(lambda x: ws['A%s' % x].value == test, [i for i in range(2, ws.max_row + 1)])[0]
    col = chr(filter(lambda x: ws['%s1' % chr(x)].value == machine, [i for i in range(66, 73)])[0])
    return "qa/utc/perform/%s/" % ws['%s%s' % (col, row)].value


def submit(test_url, values):
    length = len(values)
    test_data = {}
    constant_data = {'csrfmiddlewaretoken': token,
                     "work_started": date + ' %s' % time,
                     "work_completed": date + ' %s' % time,
                     "status": "1",
                     "form-TOTAL_FORMS": str(length),
                     "form-INITIAL_FORMS": str(length),
                     "form-MAX_NUM_FORMS": "1000"}
    for n in range(0, length):
        if values[n] == 'Yes':
            test_data["form-%s-value" % n] = '1'
        elif values[n] == 'No':
            test_data["form-%s-value" % n] = '0'
        else:
            test_data["form-%s-value" % n] = values[n]
    # test_data = {"form-%s-value" % n: values[n] for n in range(0, length)}
    submission_data = constant_data.copy()
    submission_data.update(test_data)
    resp = s.post(root + test_url, data=submission_data)


authentication_window = AuthenticationWindow(width=500, height=200, caption='please log in')

pyglet.app.run()
ip_address = read_file()
user, password, success = authentication_window.return_info()
if success:
    root, token, s = login(user, password)
    proceed = True
    date = datetime.datetime.now().strftime("%d-%m-%Y")
    time = datetime.datetime.now().strftime("%H:%M")
else:
    exit()
if __name__ == '__main__':
    print("time")





