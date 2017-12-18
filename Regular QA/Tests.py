from openpyxl import *
import glob
path = 'Machines'
# testinfo
#   test: [sheet, [excel cells]]


class tests:
    wb = load_workbook('tests.xlsx')
    ws = wb['test_cells']
    test_info = {}
    col_count = ws.max_row

    for i in range(2, ws.max_row + 1):
        test_name = str(ws['A%s' % i].value)
        sheet_name = str(ws['B%s' % i].value)
        values = [str(cell.value) for x in ws.iter_rows(min_row=i, max_row=i, min_col=3) for cell in x]
        values = filter(lambda x: x != 'None', values)
        test_info[test_name] = [sheet_name, values]

    # print test_info

    # test_info = {'ML04': ['ML04 ML05 ML09', ['S20', 'S22', 'S24', 'S26', 'S28']],
    #
    #              'ML05': ['ML04 ML05 ML09', ['J44', 'J46', 'J48']],
    #
    #              'ML06': ['ML06 ML08 ML10', ['N22', 'N24', 'T22', 'T24']],
    #
    #              'ML07_LAT': ['ML07 - LNG, LAT, VRT', ['Q20', 'Q22', 'Q24', 'Q26']],
    #
    #              'ML07_LNG': ['ML07 - LNG, LAT, VRT', ['Q39', 'Q41', 'Q43', 'Q45']],
    #
    #              'ML07_VRT': ['ML07 - LNG, LAT, VRT', ['Q58', 'Q60', 'Q62', 'Q64']],
    #
    #              'ML08': ['ML06 ML08 ML10', ['N44', 'N46', 'T44', 'T46']],
    #
    #              'ML09': ['ML04 ML05 ML09', ['J60', 'J62', 'J64']],
    #
    #              'ML10': ['ML06 ML08 ML10', ['F63', 'F65', 'F67', 'K63', 'K65', 'K67', 'P63', 'P65', 'P67']],
    #
    #              'ML11_e': ['ML11', ['AF43', 'AJ45', 'AF47', 'AJ49', 'AF51', 'AJ53', 'AF55', 'AJ57', 'AF59', 'AJ61']],
    #              'ML11_p': ['ML11', ['AF71', 'AJ73', 'AF75', 'AJ77', 'AF79', 'AJ81', 'AF83', 'AJ85', 'AF87', 'AJ89']],
    #
    #              'ML12_e': ['ML12', ['U34', 'AE34', 'AO34', 'U37', 'AE37', 'AO37', 'U40', 'AE40', 'AO40',
    #                                  'U43', 'AE43', 'AO43', 'U46', 'AE46', 'AO46']],
    #              'ML12_p': ['ML12', ['U55', 'AE55', 'AO55', 'U58', 'AE58', 'AO58', 'U61', 'AE61', 'AO61',
    #                                  'U64', 'AE64', 'AO64', 'U67', 'AE67', 'AO67']],
    #
    #              'ML13_e': ['ML13', ['G46', 'Q46', 'AA46', 'AK46', 'G48', 'Q48', 'AA48', 'AK48',
    #                                  'G50', 'Q50', 'AA50', 'AK50', 'G52', 'Q52', 'AA52', 'AK52',
    #                                  'G54', 'Q54', 'AA54', 'AK54']],
    #              'ML13_p': ['ML13', ['G27', 'Q27', 'AA27', 'AK47', 'G29', 'Q29', 'AA29', 'AK29',
    #                                  'G31', 'Q31', 'AA31', 'AK31', 'G33', 'Q33', 'AA33', 'AK33',
    #                                  'G35', 'Q35', 'AA35', 'AK35']],
    #
    #              'ML14': ['ML14 ML15', ['H18', 'O18', 'H20', 'O20', 'H22', 'O22',
    #                                     'H24', 'O24', 'H26', 'O26', 'H28', 'O28']],
    #
    #              'ML15': ['ML14 ML15', ['U44', 'U46', 'U48', 'U50', 'U52', 'U54', 'U56', 'U58', 'U60', 'U62']],
    #
    #              'ML17_iRA_2_06X': ['ML17-18', ['J23', 'J24', 'J25']],
    #              'ML17_qMLC_06X': ['ML17-18', ['J29', 'J30', 'J31', 'J32', 'J33', 'J34', 'J35', 'J36', 'J37']],
    #              'ML17_qMLC_06XFFF': ['ML17-18', ['O29', 'O30', 'O31', 'O32', 'O33', 'O34', 'O35', 'O36', 'O37']],
    #              'ML17_qMLC_10X': ['ML17-18', ['T29', 'T30', 'T31', 'T32', 'T33', 'T34', 'T35', 'T36', 'T37']],
    #              'ML17_qMLC_10XFFF': ['ML17-18', ['Y29', 'Y30', 'Y31', 'Y32', 'Y33', 'Y34', 'Y35', 'Y36', 'Y37']],
    #            'ML17_qMLC_15X': ['ML17-18', ['AD29', 'AD30', 'AD31', 'AD32', 'AD33', 'AD34', 'AD35', 'AD36', 'AD37']],
    #
    #              'DTRR': ['DTRR, DIME', ['T27']],
    #              'DIME': ['DTRR, DIME', ['T56']],
    #              'BSIV': ['MSV BSIV', ['I53', 'P53', 'I56', 'P56', 'I59', 'P59', 'I62', 'P62']],
    #              'MSV': ['MSV BSIV', ['Q21', 'X21', 'AE21', 'Q24', 'X24', 'AE24', 'Q27', 'X27', 'AE27',
    #                                   'Q30', 'X30', 'AE30', 'Q33', 'X33', 'AE33', 'Q36', 'X36', 'AE36']],
    #              }

    def get_test(self, machine, test_number):
        test_path = '%s/%s/*' % (path, machine)
        # glob loads all files in a folder
        files = glob.glob(test_path)
        # filter to files containing ml_04
        file_name = test_number.split('_')[0]
        files = filter(lambda a: file_name in a, files)
        # check that file exists

        if not files:

            return [[machine, test_number, 'nf']]

        # check if 2 files present
        elif len(files) > 1:
            # if so remove temp files
            files = filter(lambda b: '~$' not in b, files)
        elif len(files) > 1:
            files = filter(lambda b: 'mm-dd' not in b, files)

        try:

            qa_file = load_workbook(files[0], data_only=True)
            qa_sheet = qa_file[self.test_info[test_number][0]]
        except KeyError:
            test_keys = self.test_info.keys()
            test_keys = filter(lambda x: test_number in x, test_keys)
            values_to_submit = [val for sublist in [self.get_test(machine, x) for x in test_keys] for val in sublist]
            errors = map(lambda x: type(x[2]) != list, values_to_submit)
            print errors
            if True in errors:
                return [[machine, test_number, 'ev']]
            else:
                return values_to_submit


        except:
            return [[machine, test_number, 'df']]
        values_to_submit = [qa_sheet[c].value for c in self.test_info[test_number][1]]

        if None in values_to_submit:
            return [[machine, test_number, 'ev']]
        try:
            values_to_submit = map(float, values_to_submit)
        except:
            values_to_submit = map(str, values_to_submit)

        return [[machine, test_number, values_to_submit]]


if __name__ == "__main__":
    tests = tests()
